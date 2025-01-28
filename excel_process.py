import openpyxl
import openpyxl.worksheet

class ExcelProcess(object):
    # インスタンスの初期化
    def __init__(self, worksheet: openpyxl.worksheet, min_row = 2, max_row = 10):
        self.ws = worksheet
        self.min_row = min_row
        self.max_row = max_row

    def getProductsInfo(self):
        # シートから商品情報を取得
        products_info = []
        for i in range(self.min_row, self.max_row + 1):
            # JANコードを取得
            jan = self.ws.cell(row = i, column = 1).value

            # メーカーを取得
            maker = self.ws.cell(row = i, column = 2).value
            # ブランド名を取得
            brand = self.ws.cell(row = i, column = 3).value
            # 商品名を取得
            product_name = self.ws.cell(row = i, column = 4).value
            # 検索キーワードを生成
            search_keyword = product_name.replace(" ", "").replace("　", "")
            if not brand is None:
                search_keyword = brand.replace(" ", "").replace("　", "") + " " + search_keyword
            if not maker is None:
                search_keyword = maker.replace(" ", "").replace("　", "") + " " + search_keyword

            # 税込価格を取得
            price = self.ws.cell(row = i, column = 5).value

            products_info.append({
                "jan": jan,
                "keyword": search_keyword,
                "price": price
            })

        return products_info
